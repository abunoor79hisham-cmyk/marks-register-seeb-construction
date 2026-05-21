import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import re

src_xlsx = r"E:\Construction APP\Files\Construction and Building Engineering- ATTENDANCE (3).xlsx"
out_path = r"E:\Construction APP\Marks Regester\Construction_Department_MARK_SHEET.xlsx"

# Step 1: Extract all student data from xlsx with clean names
wb_src = openpyxl.load_workbook(src_xlsx)
xlsx_students = []

for sheet_name in wb_src.sheetnames:
    ws = wb_src[sheet_name]
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        r0 = str(row[0] or "").strip()
        r1 = str(row[1] or "").strip()
        r2 = str(row[2] or "").strip()
        if r0.isdigit() and r1.isdigit() and len(r1) >= 3:
            name = r2.replace('\xa0', ' ').strip()
            name = re.sub(r'  +', ' ', name)  # collapse multiple spaces
            xlsx_students.append({"s_no": int(r0), "id": r1, "name": name, "sheet": sheet_name})

wb_src.close()

# Step 2: Parse attendance_data.js for course/group/instructor mapping
import json
with open(r"E:\Construction APP\Files\attendance_data.js", 'r', encoding='utf-8-sig') as f:
    content = f.read()
js_match = re.search(r'const\s+ATTENDANCE_DATA\s*=\s*(\[.*?\]);', content, re.DOTALL)
raw = js_match.group(1)

# Convert JS object syntax to valid JSON by quoting unquoted keys
def quote_js_keys(js_text):
    """Quote unquoted JS object keys to make valid JSON"""
    # Pattern: {key:value or {key:"value or {key:'value
    # We need to find patterns like {word: or ,word:
    result = re.sub(r'([{,]\s*)([a-zA-Z_][a-zA-Z0-9_]*)(\s*:)', r'\1"\2"\3', js_text)
    return result

raw = quote_js_keys(raw)
# Also handle trailing commas
raw = re.sub(r',\s*\]', ']', raw)
raw = re.sub(r',\s*\}', '}', raw)
# Handle escaped single quotes inside strings (like \'Abdullah)
# JSON uses \" not \'
raw = raw.replace("\\'", "'")

attendance_data = json.loads(raw)

# Group JS data by (course, group) -> list of student names + instructor
from collections import OrderedDict
js_groups = OrderedDict()
for item in attendance_data:
    course = item["course"].strip()
    group = item["group"].strip()
    instructor = item.get("instructor", "").strip()
    student = item["student"].strip()
    # Normalize student name (collapse spaces, remove extra spaces)
    student = re.sub(r'  +', ' ', student)
    key = (course, group)
    if key not in js_groups:
        js_groups[key] = {"instructor": instructor, "students": []}
    js_groups[key]["students"].append(student)

# Step 3: For each JS group, find matching students in xlsx by name
def normalize_name(n):
    """Normalize name for comparison"""
    n = n.replace('\xa0', ' ').strip()
    n = re.sub(r'  +', ' ', n)
    n = n.replace("'", "").replace('"', '').replace('-', ' ').strip()
    return n.lower()

def name_similarity(a, b):
    a_parts = set(normalize_name(a).split())
    b_parts = set(normalize_name(b).split())
    if not a_parts or not b_parts:
        return 0
    return len(a_parts & b_parts) / max(len(a_parts), len(b_parts))

# Build index of xlsx students by sheet
xlsx_by_sheet = {}
for s in xlsx_students:
    sheet = s["sheet"]
    if sheet not in xlsx_by_sheet:
        xlsx_by_sheet[sheet] = []
    xlsx_by_sheet[sheet].append(s)

# Also index by name
xlsx_by_name = {}
for s in xlsx_students:
    nm = normalize_name(s["name"])
    if nm not in xlsx_by_name:
        xlsx_by_name[nm] = []
    xlsx_by_name[nm].append(s)

# Match students from JS to xlsx
groups_final = []
unmatched_students = []

for key, gdata in js_groups.items():
    course, group = key
    instructor = gdata["instructor"]
    js_students = gdata["students"]
    
    matched_students = []
    for js_sname in js_students:
        js_norm = normalize_name(js_sname)
        best_match = None
        best_score = 0
        
        if js_norm in xlsx_by_name:
            for candidate in xlsx_by_name[js_norm]:
                score = name_similarity(js_sname, candidate["name"])
                if score > best_score:
                    best_score = score
                    best_match = candidate
        
        if not best_match or best_score < 0.5:
            # Fuzzy search across all
            for candidate in xlsx_students:
                score = name_similarity(js_sname, candidate["name"])
                if score > best_score:
                    best_score = score
                    best_match = candidate
        
        if best_match and best_score >= 0.4:
            matched_students.append(best_match)
        else:
            unmatched_students.append((course, group, js_sname))
    
    # Sort by s_no
    matched_students.sort(key=lambda x: x["s_no"])
    
    groups_final.append({
        "course": course,
        "group": group,
        "instructor": instructor,
        "students": matched_students
    })

# Step 4: Create Excel workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)

header_font = Font(name='Calibri', bold=True, size=11)
title_font = Font(name='Calibri', bold=True, size=14)
subtitle_font = Font(name='Calibri', bold=True, size=10)
data_font = Font(name='Calibri', size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
max_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
total_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
sign_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

assessments = ['Quiz 1', 'Quiz 2', 'Assignment 1', 'Assignment 2', 'Midterm']
num_assess = len(assessments)
summary_rows = []

for gf_idx, gf in enumerate(groups_final):
    students = gf["students"]
    if not students:
        continue
    
    sheet_index = gf_idx + 1
    grp_short = re.sub(r'[\\/*?\[\]:]', '', gf["group"]).replace(' ', '_')[:20]
    sheet_name = f"{sheet_index}_{grp_short}"[:31]
    ws = wb.create_sheet(title=sheet_name)
    
    course = gf["course"]
    group = gf["group"]
    instructor = gf["instructor"]
    
    summary_rows.append((sheet_name, course, group, instructor, len(students)))
    
    num_cols = 3 + num_assess + 1
    
    # Row 1: Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    c = ws.cell(1, 1, "Seeb Vocational College - Construction Department\nMARK SHEET")
    c.font = title_font; c.alignment = center_align
    
    # Row 2: Info
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    instr_display = instructor if instructor else "___________"
    c = ws.cell(2, 1, f"Subject: {course}    |    Group: {group}    |    Instructor: {instr_display}")
    c.font = subtitle_font; c.alignment = center_align
    
    # Row 3: Max marks
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)
    c = ws.cell(3, 1, "Max Marks (edit here)")
    c.font = Font(name='Calibri', bold=True, size=9, color='0000FF')
    c.alignment = center_align; c.fill = max_fill; c.border = thin_border
    
    default_max = [10, 10, 15, 15, 50]
    for a_idx, (assess, dmax) in enumerate(zip(assessments, default_max)):
        col = 4 + a_idx
        c = ws.cell(3, col, dmax)
        c.font = Font(name='Calibri', bold=True, size=10, color='0000FF')
        c.alignment = center_align; c.fill = max_fill; c.border = thin_border
    
    c = ws.cell(3, num_cols)
    c.value = f"=SUM(D3:{get_column_letter(3+num_assess)}3)"
    c.font = Font(name='Calibri', bold=True, size=10, color='0000FF')
    c.alignment = center_align; c.fill = max_fill; c.border = thin_border
    
    # Row 4: Headers
    headers = ['S#', 'ID', 'Name'] + assessments + ['Total']
    for h_idx, h in enumerate(headers):
        col = 1 + h_idx
        c = ws.cell(4, col, h)
        c.font = header_font; c.alignment = center_align
        c.fill = header_fill; c.border = thin_border
    
    # Data rows
    data_start_row = 5
    for s_idx, student in enumerate(students):
        row = data_start_row + s_idx
        
        ws.cell(row, 1, student['s_no']).font = data_font
        ws.cell(row, 1).alignment = center_align; ws.cell(row, 1).border = thin_border
        
        ws.cell(row, 2, int(student['id'])).font = data_font
        ws.cell(row, 2).alignment = center_align; ws.cell(row, 2).border = thin_border
        
        ws.cell(row, 3, student['name']).font = data_font
        ws.cell(row, 3).alignment = left_align; ws.cell(row, 3).border = thin_border
        
        for a_idx in range(num_assess):
            col = 4 + a_idx
            ws.cell(row, col, None).font = data_font
            ws.cell(row, col).alignment = center_align; ws.cell(row, col).border = thin_border
        
        assess_cols = f"D{row}:{get_column_letter(3+num_assess)}{row}"
        ws.cell(row, num_cols).value = f"=SUM({assess_cols})"
        ws.cell(row, num_cols).font = Font(name='Calibri', bold=True, size=10)
        ws.cell(row, num_cols).alignment = center_align; ws.cell(row, num_cols).border = thin_border
    
    last_data_row = data_start_row + len(students) - 1
    total_row = last_data_row + 1
    
    # Total row
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    c = ws.cell(total_row, 1, "Total Marks")
    c.font = Font(name='Calibri', bold=True, size=10)
    c.alignment = center_align; c.fill = total_fill; c.border = thin_border
    
    for a_idx in range(num_assess):
        col = 4 + a_idx
        col_letter = get_column_letter(col)
        c = ws.cell(total_row, col)
        c.value = f"=SUM({col_letter}{data_start_row}:{col_letter}{last_data_row})"
        c.font = Font(name='Calibri', bold=True, size=10)
        c.alignment = center_align; c.fill = total_fill; c.border = thin_border
    
    col_letter_total = get_column_letter(num_cols)
    c = ws.cell(total_row, num_cols)
    c.value = f"=SUM({col_letter_total}{data_start_row}:{col_letter_total}{last_data_row})"
    c.font = Font(name='Calibri', bold=True, size=10)
    c.alignment = center_align; c.fill = total_fill; c.border = thin_border
    
    # Signature footer
    sign_row = total_row + 2
    ws.merge_cells(start_row=sign_row, start_column=1, end_row=sign_row, end_column=2)
    c = ws.cell(sign_row, 1, "Student Signature:")
    c.font = Font(name='Calibri', italic=True, size=9); c.alignment = left_align; c.fill = sign_fill
    ws.merge_cells(start_row=sign_row, start_column=3, end_row=sign_row, end_column=num_cols)
    c = ws.cell(sign_row, 3, "")
    c.font = Font(name='Calibri', italic=True, size=9); c.fill = sign_fill
    c.border = Border(bottom=Side(style='thin'))
    
    sign_row2 = sign_row + 1
    ws.merge_cells(start_row=sign_row2, start_column=1, end_row=sign_row2, end_column=2)
    c = ws.cell(sign_row2, 1, "Instructor Signature:")
    c.font = Font(name='Calibri', italic=True, size=9); c.alignment = left_align; c.fill = sign_fill
    
    mid_col = int(num_cols/2) + 1
    ws.merge_cells(start_row=sign_row2, start_column=3, end_row=sign_row2, end_column=mid_col)
    c = ws.cell(sign_row2, 3, instr_display)
    c.font = Font(name='Calibri', italic=True, size=9); c.fill = sign_fill
    c.border = Border(bottom=Side(style='thin'))
    
    ws.merge_cells(start_row=sign_row2, start_column=mid_col+1, end_row=sign_row2, end_column=num_cols)
    c = ws.cell(sign_row2, mid_col+1, "H.O.D Signature:")
    c.font = Font(name='Calibri', italic=True, size=9)
    c.alignment = Alignment(horizontal='right', vertical='center'); c.fill = sign_fill
    
    sign_row3 = sign_row2 + 1
    ws.merge_cells(start_row=sign_row3, start_column=mid_col+1, end_row=sign_row3, end_column=num_cols)
    c = ws.cell(sign_row3, mid_col+1, "")
    c.fill = sign_fill; c.border = Border(bottom=Side(style='thin'))
    
    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 40
    for a_idx in range(num_assess):
        ws.column_dimensions[get_column_letter(4 + a_idx)].width = 14
    ws.column_dimensions[get_column_letter(num_cols)].width = 10

# --- Index sheet ---
ws_idx = wb.create_sheet(title="Index", index=0)
ws_idx.merge_cells('A1:E1')
c = ws_idx.cell(1, 1, "Construction Department - Mark Sheet Index")
c.font = Font(name='Calibri', bold=True, size=16); c.alignment = center_align

for i, h in enumerate(['Sheet Name', 'Subject', 'Group', 'Instructor', 'No. of Students']):
    col = i + 1
    c = ws_idx.cell(2, col, h)
    c.font = Font(name='Calibri', bold=True, size=10)
    c.alignment = center_align; c.fill = header_fill; c.border = thin_border

for r_idx, (sname, subj, grp, instr, count) in enumerate(summary_rows):
    row = 3 + r_idx
    vals = [sname, subj, grp, instr, count]
    for c_idx, v in enumerate(vals):
        c = ws_idx.cell(row, c_idx + 1, v)
        c.font = Font(name='Calibri', size=10)
        c.alignment = center_align if c_idx != 1 else left_align
        c.border = thin_border

ws_idx.column_dimensions['A'].width = 32
ws_idx.column_dimensions['B'].width = 45
ws_idx.column_dimensions['C'].width = 25
ws_idx.column_dimensions['D'].width = 25
ws_idx.column_dimensions['E'].width = 18

wb.save(out_path)
print(f"Created: {out_path}")
print(f"Sheets: {len(wb.sheetnames)} (Index + {len(wb.sheetnames)-1} groups)")
print(f"Groups processed: {len(summary_rows)}")
if unmatched_students:
    print(f"\nUnmatched students ({len(unmatched_students)}):")
    for c, g, s in unmatched_students[:20]:
        print(f"  {c} / {g}: {s}")
