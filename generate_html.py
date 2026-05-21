import json, re
from collections import OrderedDict

# --- Load data from JS file ---
with open(r"E:\Construction APP\Files\attendance_data.js", 'r', encoding='utf-8-sig') as f:
    content = f.read()
js_match = re.search(r'const\s+ATTENDANCE_DATA\s*=\s*(\[.*?\]);', content, re.DOTALL)
raw = js_match.group(1)
raw = re.sub(r'([{,]\s*)([a-zA-Z_][a-zA-Z0-9_]*)(\s*:)', r'\1"\2"\3', raw)
raw = re.sub(r',\s*\]', ']', raw)
raw = re.sub(r',\s*\}', '}', raw)
raw = raw.replace("\\'", "'")
attendance_data = json.loads(raw)

# --- Load student IDs from xlsx ---
import openpyxl
wb_src = openpyxl.load_workbook(r"E:\Construction APP\Files\Construction and Building Engineering- ATTENDANCE (3).xlsx")
xlsx_students = {}
for sheet_name in wb_src.sheetnames:
    ws = wb_src[sheet_name]
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        r0 = str(row[0] or "").strip(); r1 = str(row[1] or "").strip(); r2 = str(row[2] or "").strip()
        if r0.isdigit() and r1.isdigit() and len(r1) >= 3:
            name = r2.replace('\xa0', ' ').strip(); name = re.sub(r'  +', ' ', name).strip().lower()
            xlsx_students[name] = {"s_no": int(r0), "id": r1}
wb_src.close()

# --- Build groups ---
js_groups = OrderedDict()
for item in attendance_data:
    course = item["course"].strip(); group = item["group"].strip()
    instructor = item.get("instructor", "").strip(); student = item["student"].strip()
    student = re.sub(r'  +', ' ', student)
    key = (course, group)
    if key not in js_groups:
        js_groups[key] = {"instructor": instructor, "students": []}
    js_groups[key]["students"].append(student)

def normalize_name(n):
    n = n.replace('\xa0', ' ').strip(); n = re.sub(r'  +', ' ', n)
    n = n.replace("'", "").replace('"', '').replace('-', ' ').strip()
    return n.lower()

def name_similarity(a, b):
    a_parts = set(normalize_name(a).split()); b_parts = set(normalize_name(b).split())
    if not a_parts or not b_parts: return 0
    return len(a_parts & b_parts) / max(len(a_parts), len(b_parts))

groups_data_js = []
for key, gdata in js_groups.items():
    course, group = key; instructor = gdata["instructor"]; js_students = gdata["students"]
    matched = []
    for s in js_students:
        sn = normalize_name(s)
        best = None; best_score = 0
        for xn, xd in xlsx_students.items():
            score = name_similarity(s, xn)
            if score > best_score: best_score = score; best = xd
        if best and best_score >= 0.4:
            matched.append((best["s_no"], best["id"], s))
    matched.sort(key=lambda x: x[0])
    groups_data_js.append({"course": course, "group": group, "instructor": instructor, "students": matched})

instructors = sorted(set(g["instructor"] for g in groups_data_js if g["instructor"]))

DATA_JSON = json.dumps(groups_data_js, ensure_ascii=False)
INST_JSON = json.dumps(instructors, ensure_ascii=False)

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Construction Department - Mark Sheet System</title>
<style>
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: 'Segoe UI', Calibri, Arial, sans-serif; background: #f0f2f5; }
.nav { background: linear-gradient(135deg, #1a3a5c, #2a5f8f); color: white; padding: 0 30px; display: flex; align-items: center; height: 52px; }
.nav h1 { font-size: 16px; margin-right: 40px; white-space: nowrap; }
.nav-tabs { display: flex; gap: 2px; height: 100%; }
.nav-tab { padding: 0 20px; display: flex; align-items: center; cursor: pointer; font-size: 13px; color: rgba(255,255,255,0.75); border-bottom: 3px solid transparent; transition: all 0.15s; user-select: none; }
.nav-tab:hover { color: #fff; background: rgba(255,255,255,0.08); }
.nav-tab.active { color: #fff; border-bottom-color: #4fc3f7; background: rgba(255,255,255,0.12); font-weight: 600; }
.page { display: none; padding: 20px 30px; }
.page.active { display: block; }
.sheets-layout { display: flex; gap: 20px; }
.sheets-sidebar { width: 300px; flex-shrink: 0; background: #fff; border-radius: 8px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); overflow: hidden; display: flex; flex-direction: column; max-height: calc(100vh - 100px); }
.sheets-sidebar .sb-header { padding: 14px 16px 8px; border-bottom: 1px solid #eee; }
.sheets-sidebar .sb-header h3 { font-size: 13px; color: #555; font-weight: 600; text-transform: uppercase; }
.sheets-sidebar .sb-header select, .sheets-sidebar .sb-header input { width: 100%; margin-top: 6px; padding: 6px 8px; border: 1px solid #ccc; border-radius: 5px; font-size: 13px; }
.sheets-sidebar .sb-list { flex: 1; overflow-y: auto; }
.group-item { padding: 10px 16px; cursor: pointer; border-bottom: 1px solid #eee; font-size: 13px; transition: background 0.15s; }
.group-item:hover { background: #e8f0fe; }
.group-item.active { background: #d2e3fc; border-left: 3px solid #1a73e8; font-weight: 600; }
.group-item .badge { float: right; background: #ddd; color: #555; border-radius: 10px; padding: 1px 8px; font-size: 11px; }
.sheets-main { flex: 1; overflow-x: auto; }
.sheet-title { font-size: 18px; font-weight: 600; color: #1a3a5c; margin-bottom: 4px; }
.sheet-subtitle { font-size: 13px; color: #666; margin-bottom: 14px; }
.sheet-controls { margin-bottom: 10px; display: flex; gap: 6px; flex-wrap: wrap; }
.sheet-controls button { padding: 6px 16px; border: 1px solid #1a73e8; background: #fff; color: #1a73e8; border-radius: 4px; cursor: pointer; font-size: 12px; }
.sheet-controls button:hover { background: #e8f0fe; }
.sheet-controls button.danger { border-color: #d93025; color: #d93025; }
.sheet-controls button.danger:hover { background: #fce8e6; }
table { width: 100%; border-collapse: collapse; background: #fff; box-shadow: 0 1px 3px rgba(0,0,0,0.1); border-radius: 6px; overflow: hidden; }
th, td { padding: 6px 8px; text-align: center; border: 1px solid #d0d0d0; font-size: 13px; }
th { background: #d9e1f2; font-weight: 600; color: #1a3a5c; position: relative; }
.max-row td { background: #e2efda; color: #0066cc; font-weight: 600; font-size: 12px; }
.max-row td.max-label { color: #555; font-weight: 400; font-size: 12px; }
.max-row .mark-input { color: #0066cc; font-weight: 700; }
.total-row td { background: #fce4d6; font-weight: 600; }
.mark-input { width: 55px; padding: 4px 4px; text-align: center; border: 1px solid #ccc; border-radius: 3px; font-size: 13px; }
.mark-input:focus { outline: 2px solid #1a73e8; border-color: transparent; }
.total-cell { font-weight: 700; color: #1a3a5c; }
.footer-section { margin-top: 20px; padding: 14px 20px; background: #f7f7f7; border-radius: 6px; font-size: 13px; }
.footer-section .sig-row { display: flex; justify-content: space-between; margin-bottom: 6px; }
.footer-section .sig-row span { color: #555; }
.footer-section .sig-line { display: inline-block; min-width: 180px; border-bottom: 1px solid #333; margin-left: 4px; height: 18px; }
.dashboard-stats { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 16px; margin-bottom: 24px; }
.stat-card { background: #fff; border-radius: 8px; padding: 16px 20px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }
.stat-card .stat-num { font-size: 28px; font-weight: 700; color: #1a3a5c; }
.stat-card .stat-label { font-size: 12px; color: #777; margin-top: 2px; }
.progress-container { background: #fff; border-radius: 8px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); padding: 16px 20px; margin-bottom: 12px; }
.progress-group { display: flex; align-items: center; padding: 8px 0; border-bottom: 1px solid #f0f0f0; }
.progress-group:last-child { border-bottom: none; }
.progress-group .pg-info { width: 280px; flex-shrink: 0; }
.progress-group .pg-info strong { font-size: 13px; }
.progress-group .pg-info small { display: block; color: #888; font-size: 11px; }
.progress-bar-wrap { flex: 1; height: 22px; background: #e9ecef; border-radius: 11px; overflow: hidden; margin: 0 16px; position: relative; }
.progress-bar-fill { height: 100%; border-radius: 11px; transition: width 0.5s ease; background: linear-gradient(90deg, #4caf50, #8bc34a); }
.progress-bar-fill.low { background: linear-gradient(90deg, #f44336, #ff7043); }
.progress-bar-fill.medium { background: linear-gradient(90deg, #ff9800, #ffc107); }
.progress-bar-fill.good { background: linear-gradient(90deg, #2196f3, #03a9f4); }
.progress-pct { width: 50px; text-align: right; font-weight: 600; font-size: 13px; color: #333; flex-shrink: 0; }
.analysis-select { margin-bottom: 16px; }
.analysis-select select { padding: 6px 12px; border: 1px solid #ccc; border-radius: 5px; font-size: 14px; min-width: 300px; }
.analysis-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-bottom: 20px; }
.analysis-card { background: #fff; border-radius: 8px; padding: 16px 20px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }
.analysis-card h4 { font-size: 13px; color: #555; font-weight: 600; margin-bottom: 10px; text-transform: uppercase; }
.stat-row { display: flex; justify-content: space-between; padding: 4px 0; font-size: 13px; border-bottom: 1px solid #f5f5f5; }
.stat-row .stat-val { font-weight: 600; }
.highlight-green { color: #2e7d32; font-weight: 700; }
.highlight-red { color: #c62828; font-weight: 700; }
.highlight-orange { color: #e65100; font-weight: 700; }
.admin-login { max-width: 400px; margin: 40px auto; background: #fff; border-radius: 8px; padding: 30px; box-shadow: 0 2px 8px rgba(0,0,0,0.12); text-align: center; }
.admin-login h2 { margin-bottom: 20px; color: #1a3a5c; }
.admin-login input { width: 100%; padding: 10px 14px; margin-bottom: 12px; border: 1px solid #ccc; border-radius: 5px; font-size: 14px; }
.admin-login button { width: 100%; padding: 10px; background: #1a73e8; color: #fff; border: none; border-radius: 5px; font-size: 14px; cursor: pointer; }
.admin-login button:hover { background: #1557b0; }
.admin-login .error { color: #d93025; font-size: 13px; margin-top: 8px; display: none; }
.admin-panel { max-width: 600px; margin: 20px auto; }
.admin-card { background: #fff; border-radius: 8px; padding: 24px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); margin-bottom: 16px; }
.admin-card h3 { font-size: 15px; color: #1a3a5c; margin-bottom: 12px; }
.admin-card p { font-size: 13px; color: #666; margin-bottom: 12px; }
.admin-card input[type=file] { display: block; margin-bottom: 10px; font-size: 13px; }
.admin-card button { padding: 8px 20px; border: 1px solid #1a73e8; background: #1a73e8; color: #fff; border-radius: 4px; cursor: pointer; font-size: 13px; }
.admin-card button:hover { background: #1557b0; }
.admin-card button.secondary { background: #fff; color: #1a73e8; }
.admin-card .msg { margin-top: 8px; font-size: 12px; color: #2e7d32; display: none; }
.no-sheet { text-align: center; padding: 60px; color: #999; font-size: 16px; }
@media print { .nav, .sheets-sidebar, .sheet-controls, .no-print { display: none !important; } }
</style>
</head>
<body>

<div class="nav">
  <h1>Seeb VC - Construction</h1>
  <div class="nav-tabs">
    <div class="nav-tab active" data-tab="sheets" onclick="switchTab('sheets')">Mark Sheets</div>
    <div class="nav-tab" data-tab="dashboard" onclick="switchTab('dashboard')">Dashboard</div>
    <div class="nav-tab" data-tab="analysis" onclick="switchTab('analysis')">Analysis</div>
    <div class="nav-tab" data-tab="admin" onclick="switchTab('admin')">Admin</div>
  </div>
</div>

<!-- MARKS SHEETS -->
<div id="page-sheets" class="page active">
  <div class="sheets-layout">
    <div class="sheets-sidebar">
      <div class="sb-header">
        <h3>Groups</h3>
        <select id="instructorFilter" onchange="filterByInstructor()"><option value="">All Instructors</option></select>
        <input type="text" id="searchInput" placeholder="Search groups..." onkeyup="filterGroups()">
      </div>
      <div class="sb-list" id="groupList"></div>
    </div>
    <div class="sheets-main" id="mainContent"><div class="no-sheet">Select a group from the sidebar</div></div>
  </div>
</div>

<!-- DASHBOARD -->
<div id="page-dashboard" class="page">
  <div class="dashboard-stats" id="dashStats"></div>
  <div id="dashProgress"></div>
</div>

<!-- ANALYSIS -->
<div id="page-analysis" class="page">
  <div class="analysis-select">
    <select id="analysisGroupSelect" onchange="renderAnalysis()"><option value="">-- Select a group --</option></select>
  </div>
  <div id="analysisContent"><div class="no-sheet">Select a group above to view analysis</div></div>
</div>

<!-- ADMIN -->
<div id="page-admin" class="page">
  <div id="adminLogin" class="admin-login">
    <h2>Admin Login</h2>
    <input type="text" id="adminUser" placeholder="Email">
    <input type="password" id="adminPass" placeholder="Password">
    <button onclick="adminLogin()">Sign In</button>
    <div class="error" id="adminError">Invalid credentials</div>
  </div>
  <div id="adminPanel" class="admin-panel" style="display:none">
    <div class="admin-card">
      <h3>Upload New Database</h3>
      <p>Upload a JSON file containing updated group/course data to replace the current database.</p>
      <input type="file" id="uploadDbInput" accept=".json">
      <button onclick="uploadDatabase()">Upload & Replace</button>
      <div class="msg" id="uploadMsg">Database updated successfully!</div>
    </div>
    <div class="admin-card">
      <h3>Save Current Marks</h3>
      <p>Download all entered marks data as a JSON backup file.</p>
      <button onclick="saveMarks()">Download Marks Backup</button>
    </div>
    <div class="admin-card">
      <h3>Load Marks Backup</h3>
      <p>Restore marks from a previously saved backup JSON file.</p>
      <input type="file" id="loadMarksInput" accept=".json">
      <button onclick="loadMarks()">Restore Marks</button>
      <div class="msg" id="loadMsg">Marks restored successfully!</div>
    </div>
    <div class="admin-card">
      <h3>Sign Out</h3>
      <p>You are signed in as <span id="adminEmailDisplay"></span></p>
      <button class="secondary" onclick="adminLogout()">Sign Out</button>
    </div>
  </div>
</div>

<script>
window.onerror = function(msg, url, line) {
  document.body.innerHTML = '<div style="color:red;padding:30px;font-size:18px"><h2>JavaScript Error</h2><p>'+msg+'</p><p>Line: '+line+'</p><p>URL: '+url+'</p></div>';
  return true;
};
// DATA INJECTION
const GROUPS_DATA = __DATA_JSON__;
const INSTRUCTORS = __INST_JSON__;

let currentIndex = -1;
let marksData = {};
let filteredIndices = [];

function init() {
  loadFromStorage();
  populateInstructorFilter();
  populateAnalysisSelect();
  buildGroupList();
  renderDashboard();
  if (GROUPS_DATA.length > 0) selectGroup(0);
}

function saveToStorage() {
  try { localStorage.setItem('marksData', JSON.stringify(marksData)); } catch(e) {}
}

function loadFromStorage() {
  try {
    const saved = localStorage.getItem('marksData');
    if (saved) { const parsed = JSON.parse(saved); if (typeof parsed === 'object' && !Array.isArray(parsed)) { marksData = parsed; return; } }
  } catch(e) {}
  initMarksData();
}

function initMarksData() {
  marksData = {};
  GROUPS_DATA.forEach((g, i) => {
    marksData[i] = {};
    g.students.forEach((s, si) => {
      marksData[i][si+'_q1'] = ''; marksData[i][si+'_q2'] = '';
      marksData[i][si+'_a1'] = ''; marksData[i][si+'_a2'] = '';
      marksData[i][si+'_mid'] = '';
    });
    marksData[i]['max_q1'] = 10; marksData[i]['max_q2'] = 10;
    marksData[i]['max_a1'] = 15; marksData[i]['max_a2'] = 15; marksData[i]['max_mid'] = 50;
  });
}

// Tab switching
let currentTab = 'sheets';
function switchTab(tab) {
  currentTab = tab;
  document.querySelectorAll('.page').forEach(p => p.classList.remove('active'));
  document.querySelectorAll('.nav-tab').forEach(t => t.classList.remove('active'));
  document.getElementById('page-'+tab).classList.add('active');
  document.querySelector('.nav-tab[data-tab="'+tab+'"]').classList.add('active');
  if (tab === 'dashboard') renderDashboard();
  if (tab === 'analysis') renderAnalysis();
}

// Instructor filter
function populateInstructorFilter() {
  const sel = document.getElementById('instructorFilter');
  INSTRUCTORS.forEach(inst => { const o = document.createElement('option'); o.value = inst; o.textContent = inst; sel.appendChild(o); });
}

function filterByInstructor() { buildGroupList(); }
function filterGroups() { buildGroupList(); }

function buildGroupList() {
  const list = document.getElementById('groupList');
  const searchQ = document.getElementById('searchInput').value.toLowerCase();
  const instFilter = document.getElementById('instructorFilter').value;
  list.innerHTML = '';
  filteredIndices = [];
  GROUPS_DATA.forEach((g, i) => {
    if (instFilter && g.instructor !== instFilter) return;
    const searchText = (g.group+' '+g.course+' '+g.instructor).toLowerCase();
    if (searchQ && !searchText.includes(searchQ)) return;
    filteredIndices.push(i);
    const div = document.createElement('div');
    div.className = 'group-item';
    div.dataset.index = i;
    div.innerHTML = '<strong>'+escHtml(g.group)+'</strong><br><small>'+escHtml(g.course)+'</small><span class="badge">'+g.students.length+'</span>';
    div.onclick = function() { selectGroup(i); };
    list.appendChild(div);
  });
  if (filteredIndices.includes(currentIndex)) {
    document.querySelectorAll('.group-item').forEach(el => { if (parseInt(el.dataset.index) === currentIndex) el.classList.add('active'); });
  }
}

// ============ MARKS SHEETS ============
function selectGroup(idx) {
  currentIndex = idx;
  document.querySelectorAll('.group-item').forEach(el => el.classList.remove('active'));
  document.querySelectorAll('.group-item').forEach(el => { if (parseInt(el.dataset.index) === idx) el.classList.add('active'); });
  renderSheet(idx);
}

function renderSheet(idx) {
  const g = GROUPS_DATA[idx];
  const main = document.getElementById('mainContent');
  const m = marksData[idx] || {};
  let html = '<div class="sheet-title">'+escHtml(g.course)+'</div>';
  html += '<div class="sheet-subtitle">Group: <strong>'+escHtml(g.group)+'</strong> &nbsp;|&nbsp; Instructor: <strong>'+escHtml(g.instructor||'___________')+'</strong></div>';
  html += '<div class="sheet-controls no-print">';
  html += '<button onclick="exportCSV('+idx+')">Export CSV</button>';
  html += '<button onclick="window.print()">Print</button>';
  html += '<button class="danger" onclick="resetMarks('+idx+')">Clear Marks</button></div>';

  html += '<table><thead><tr>';
  html += '<th style="width:38px">S#</th><th style="width:85px">ID</th><th style="width:280px">Name</th>';
  html += '<th>Quiz 1</th><th>Quiz 2</th><th>Assignment 1</th><th>Assignment 2</th><th>Midterm</th><th style="width:65px">Total</th>';
  html += '</tr></thead><tbody>';

  // Max marks row
  html += '<tr class="max-row"><td colspan="3" class="max-label">Max Marks (editable below)</td>';
  html += '<td><input class="mark-input" type="number" min="0" max="200" value="'+(m.max_q1||10)+'" onchange="updateMax('+idx+',\'q1\',this.value)"></td>';
  html += '<td><input class="mark-input" type="number" min="0" max="200" value="'+(m.max_q2||10)+'" onchange="updateMax('+idx+',\'q2\',this.value)"></td>';
  html += '<td><input class="mark-input" type="number" min="0" max="200" value="'+(m.max_a1||15)+'" onchange="updateMax('+idx+',\'a1\',this.value)"></td>';
  html += '<td><input class="mark-input" type="number" min="0" max="200" value="'+(m.max_a2||15)+'" onchange="updateMax('+idx+',\'a2\',this.value)"></td>';
  html += '<td><input class="mark-input" type="number" min="0" max="200" value="'+(m.max_mid||50)+'" onchange="updateMax('+idx+',\'mid\',this.value)"></td>';
  var maxTotal = (parseFloat(m.max_q1)||10)+(parseFloat(m.max_q2)||10)+(parseFloat(m.max_a1)||15)+(parseFloat(m.max_a2)||15)+(parseFloat(m.max_mid)||50);
  html += '<td class="total-cell">'+maxTotal+'</td></tr>';

  // Student rows
  var fields = ['q1','q2','a1','a2','mid'];
  var totals = {q1:0,q2:0,a1:0,a2:0,mid:0,all:0};

  g.students.forEach(function(s, si) {
    var vals = fields.map(function(f) { return parseFloat(m[si+'_'+f]) || 0; });
    var total = vals[0]+vals[1]+vals[2]+vals[3]+vals[4];
    totals.q1 += vals[0]; totals.q2 += vals[1]; totals.a1 += vals[2]; totals.a2 += vals[3]; totals.mid += vals[4]; totals.all += total;
    html += '<tr><td>'+s[0]+'</td><td>'+s[1]+'</td><td class="name-col">'+escHtml(s[2])+'</td>';
    fields.forEach(function(f, fi) {
      html += '<td><input class="mark-input" type="number" min="0" max="200" value="'+(m[si+'_'+f]||'')+'" onchange="updateMark('+idx+','+si+',\''+f+'\',this.value)"></td>';
    });
    html += '<td class="total-cell">'+total+'</td></tr>';
  });

  // Totals row
  html += '<tr class="total-row"><td colspan="3">Total Marks</td>';
  html += '<td>'+totals.q1+'</td><td>'+totals.q2+'</td><td>'+totals.a1+'</td><td>'+totals.a2+'</td><td>'+totals.mid+'</td><td>'+totals.all+'</td></tr>';
  html += '</tbody></table>';

  // Footer
  html += '<div class="footer-section">';
  html += '<div class="sig-row"><span>Student Signature: <span class="sig-line"></span></span><span>Date: ___________</span></div>';
  html += '<div class="sig-row"><span>Instructor Signature: <span class="sig-line">'+escHtml(g.instructor)+'</span></span><span>H.O.D Signature: <span class="sig-line"></span></span></div>';
  html += '</div>';

  main.innerHTML = html;
}

function updateMark(idx, si, field, val) {
  marksData[idx][si+'_'+field] = val;
  saveToStorage();
  renderSheet(idx);
}

function updateMax(idx, field, val) {
  marksData[idx]['max_'+field] = val;
  saveToStorage();
  renderSheet(idx);
}

function resetMarks(idx) {
  if (!confirm('Clear all marks for this group?')) return;
  var g = GROUPS_DATA[idx];
  g.students.forEach(function(s, si) {
    marksData[idx][si+'_q1'] = ''; marksData[idx][si+'_q2'] = '';
    marksData[idx][si+'_a1'] = ''; marksData[idx][si+'_a2'] = ''; marksData[idx][si+'_mid'] = '';
  });
  saveToStorage();
  renderSheet(idx);
}

function exportCSV(idx) {
  var g = GROUPS_DATA[idx]; var m = marksData[idx];
  var csv = 'S#,ID,Name,Quiz 1,Quiz 2,Assignment 1,Assignment 2,Midterm,Total\\n';
  g.students.forEach(function(s, si) {
    var vals = ['q1','q2','a1','a2','mid'].map(function(f) { return parseFloat(m[si+'_'+f]) || 0; });
    var total = vals[0]+vals[1]+vals[2]+vals[3]+vals[4];
    csv += s[0]+','+s[1]+',"'+s[2]+'",'+vals[0]+','+vals[1]+','+vals[2]+','+vals[3]+','+vals[4]+','+total+'\\n';
  });
  var blob = new Blob([csv], { type: 'text/csv' });
  var url = URL.createObjectURL(blob);
  var a = document.createElement('a'); a.href = url;
  a.download = g.group.replace(/[^a-z0-9]/gi,'_')+'_marks.csv';
  a.click(); URL.revokeObjectURL(url);
}

// ============ DASHBOARD ============
function renderDashboard() {
  var totalGroups = GROUPS_DATA.length;
  var totalStudents = 0, totalFilled = 0, totalCells = 0;
  var groupStats = [];
  var fields = ['q1','q2','a1','a2','mid'];

  GROUPS_DATA.forEach(function(g, i) {
    var m = marksData[i] || {};
    var filled = 0; var total = g.students.length * 5;
    g.students.forEach(function(s, si) {
      fields.forEach(function(f) { var v = m[si+'_'+f]; if (v !== undefined && v !== null && v !== '') filled++; });
    });
    var pct = total > 0 ? Math.round(filled/total*100) : 0;
    totalStudents += g.students.length; totalFilled += filled; totalCells += total;
    groupStats.push({i:i, g:g, filled:filled, total:total, pct:pct});
  });

  var overallPct = totalCells > 0 ? Math.round(totalFilled/totalCells*100) : 0;

  document.getElementById('dashStats').innerHTML =
    '<div class="stat-card"><div class="stat-num">'+totalGroups+'</div><div class="stat-label">Total Groups</div></div>'+
    '<div class="stat-card"><div class="stat-num">'+totalStudents+'</div><div class="stat-label">Total Students</div></div>'+
    '<div class="stat-card"><div class="stat-num">'+totalFilled+' / '+totalCells+'</div><div class="stat-label">Marks Entered</div></div>'+
    '<div class="stat-card"><div class="stat-num">'+overallPct+'%</div><div class="stat-label">Overall Completion</div></div>';

  var phtml = '<h3 style="margin-bottom:12px;color:#1a3a5c">Progress per Group</h3>';
  groupStats.sort(function(a,b) { return a.pct - b.pct; });
  groupStats.forEach(function(ps) {
    var cls = ps.pct < 25 ? 'low' : ps.pct < 50 ? 'medium' : ps.pct < 75 ? 'good' : '';
    phtml += '<div class="progress-container"><div class="progress-group">'+
      '<div class="pg-info"><strong>'+escHtml(ps.g.group)+'</strong><small>'+escHtml(ps.g.course)+'</small></div>'+
      '<div class="progress-bar-wrap"><div class="progress-bar-fill '+cls+'" style="width:'+ps.pct+'%"></div></div>'+
      '<div class="progress-pct">'+ps.pct+'%</div></div></div>';
  });
  document.getElementById('dashProgress').innerHTML = phtml;
}

// ============ ANALYSIS ============
function populateAnalysisSelect() {
  var sel = document.getElementById('analysisGroupSelect');
  GROUPS_DATA.forEach(function(g, i) {
    var o = document.createElement('option');
    o.value = i; o.textContent = g.group+' - '+g.course+' ('+g.instructor+')';
    sel.appendChild(o);
  });
}

function renderAnalysis() {
  var sel = document.getElementById('analysisGroupSelect');
  var idx = parseInt(sel.value);
  var container = document.getElementById('analysisContent');
  if (isNaN(idx)) { container.innerHTML = '<div class="no-sheet">Select a group above to view analysis</div>'; return; }

  var g = GROUPS_DATA[idx]; var m = marksData[idx] || {};
  var fields = ['q1','q2','a1','a2','mid'];
  var fieldLabels = ['Quiz 1','Quiz 2','Assignment 1','Assignment 2','Midterm'];

  var maxDefaults = {q1:10,q2:10,a1:15,a2:15,mid:50};
  var maxVals = fields.map(function(f) { return parseFloat(m['max_'+f]) || maxDefaults[f]; });
  var maxTotal = maxVals[0]+maxVals[1]+maxVals[2]+maxVals[3]+maxVals[4];

  var studentData = [];
  g.students.forEach(function(s, si) {
    var scores = fields.map(function(f) { return parseFloat(m[si+'_'+f]) || 0; });
    var total = scores[0]+scores[1]+scores[2]+scores[3]+scores[4];
    studentData.push({s:s, si:si, scores:scores, total:total});
  });

  var avg = fields.map(function(_, fi) {
    var vals = studentData.map(function(sd) { return sd.scores[fi]; }).filter(function(v) { return v > 0; });
    return vals.length > 0 ? (vals.reduce(function(a,b){return a+b;},0)/vals.length) : 0;
  });

  var filled = 0, cellTotal = g.students.length * 5;
  g.students.forEach(function(s, si) {
    fields.forEach(function(f) { var v = m[si+'_'+f]; if (v !== undefined && v !== null && v !== '') filled++; });
  });
  var completionPct = cellTotal > 0 ? Math.round(filled/cellTotal*100) : 0;

  var halfTotal = maxTotal / 2;
  var aboveHalf = studentData.filter(function(sd) { return sd.total >= halfTotal; }).length;
  var belowHalf = studentData.filter(function(sd) { return sd.total > 0 && sd.total < halfTotal; }).length;
  var noMarks = studentData.filter(function(sd) { return sd.total === 0; }).length;

  var html = '<div class="sheet-title">'+escHtml(g.course)+'</div>';
  html += '<div class="sheet-subtitle">Group: <strong>'+escHtml(g.group)+'</strong> | Instructor: <strong>'+escHtml(g.instructor)+'</strong></div>';

  html += '<div class="analysis-grid">';
  html += '<div class="analysis-card"><h4>Completion</h4><div class="stat-row"><span>Marks Entered</span><span class="stat-val">'+filled+' / '+cellTotal+'</span></div><div class="stat-row"><span>Completion Rate</span><span class="stat-val '+(completionPct>=75?'highlight-green':completionPct>=50?'highlight-orange':'highlight-red')+'">'+completionPct+'%</span></div></div>';
  html += '<div class="analysis-card"><h4>Student Performance</h4><div class="stat-row"><span>Above 50%</span><span class="stat-val highlight-green">'+aboveHalf+'</span></div><div class="stat-row"><span>Below 50%</span><span class="stat-val highlight-orange">'+belowHalf+'</span></div><div class="stat-row"><span>No Marks</span><span class="stat-val highlight-red">'+noMarks+'</span></div></div>';

  fieldLabels.forEach(function(label, fi) {
    var maxVal = maxVals[fi]; var half = maxVal/2;
    var above = studentData.filter(function(sd) { return sd.scores[fi] >= half && sd.scores[fi] > 0; }).length;
    var zero = studentData.filter(function(sd) { return sd.scores[fi] === 0 || sd.scores[fi] === null; }).length;
    html += '<div class="analysis-card"><h4>'+label+' (Max: '+maxVal+')</h4>';
    html += '<div class="stat-row"><span>Class Average</span><span class="stat-val">'+avg[fi].toFixed(1)+'</span></div>';
    html += '<div class="stat-row"><span>Above 50% max</span><span class="stat-val highlight-green">'+above+'</span></div>';
    html += '<div class="stat-row"><span>Not entered</span><span class="stat-val highlight-red">'+zero+'</span></div></div>';
  });
  html += '</div>';

  html += '<table><thead><tr><th>S#</th><th>ID</th><th>Name</th>';
  fieldLabels.forEach(function(l) { html += '<th>'+l+'</th>'; });
  html += '<th>Total</th><th>Status</th></tr></thead><tbody>';

  studentData.forEach(function(sd) {
    var pct = maxTotal > 0 ? Math.round(sd.total/maxTotal*100) : 0;
    var status = pct>=75?'Excellent':pct>=50?'Good':pct>=25?'Weak':sd.total===0?'No marks':'Poor';
    var cls = pct>=75?'highlight-green':pct>=50?'highlight-orange':'highlight-red';
    html += '<tr><td>'+sd.s[0]+'</td><td>'+sd.s[1]+'</td><td class="name-col">'+escHtml(sd.s[2])+'</td>';
    sd.scores.forEach(function(s) { html += '<td>'+(s||'')+'</td>'; });
    html += '<td><strong>'+sd.total+'</strong></td><td class="'+cls+'" style="font-size:12px;font-weight:600">'+status+'</td></tr>';
  });
  html += '</tbody></table>';

  container.innerHTML = html;
}

// ============ ADMIN ============
const ADMIN_USER = 's6343@seebvc.edu.om';
const ADMIN_PASS = 'Abuaysam***79';

function adminLogin() {
  var u = document.getElementById('adminUser').value;
  var p = document.getElementById('adminPass').value;
  if (u === ADMIN_USER && p === ADMIN_PASS) {
    document.getElementById('adminLogin').style.display = 'none';
    document.getElementById('adminPanel').style.display = 'block';
    document.getElementById('adminEmailDisplay').textContent = ADMIN_USER;
    document.getElementById('adminError').style.display = 'none';
  } else {
    document.getElementById('adminError').style.display = 'block';
  }
}

function adminLogout() {
  document.getElementById('adminLogin').style.display = 'block';
  document.getElementById('adminPanel').style.display = 'none';
  document.getElementById('adminUser').value = '';
  document.getElementById('adminPass').value = '';
}

function uploadDatabase() {
  var fi = document.getElementById('uploadDbInput');
  var file = fi.files[0];
  if (!file) { alert('Please select a JSON file first.'); return; }
  var reader = new FileReader();
  reader.onload = function(e) {
    try {
      var data = JSON.parse(e.target.result);
      if (!Array.isArray(data)) throw new Error('Not an array');
      data.forEach(function(item, i) { if (!item.course || !item.group || !item.students) throw new Error('Item '+i+' missing required fields'); });
      localStorage.setItem('uploadedDatabase', JSON.stringify(data));
      document.getElementById('uploadMsg').style.display = 'block';
      document.getElementById('uploadMsg').textContent = 'Database updated! Reloading page...';
      setTimeout(function() { location.reload(); }, 1500);
    } catch(err) { alert('Invalid JSON file: '+err.message); }
  };
  reader.readAsText(file);
}

function saveMarks() {
  var blob = new Blob([JSON.stringify(marksData, null, 2)], { type: 'application/json' });
  var url = URL.createObjectURL(blob);
  var a = document.createElement('a'); a.href = url;
  a.download = 'marks_backup_'+new Date().toISOString().slice(0,10)+'.json';
  a.click(); URL.revokeObjectURL(url);
}

function loadMarks() {
  var fi = document.getElementById('loadMarksInput');
  var file = fi.files[0];
  if (!file) { alert('Please select a backup JSON file.'); return; }
  var reader = new FileReader();
  reader.onload = function(e) {
    try {
      var data = JSON.parse(e.target.result);
      if (typeof data !== 'object' || Array.isArray(data)) throw new Error('Invalid marks data format');
      marksData = data;
      saveToStorage();
      document.getElementById('loadMsg').style.display = 'block';
      if (currentIndex >= 0) renderSheet(currentIndex);
      if (currentTab === 'dashboard') renderDashboard();
      if (currentTab === 'analysis') renderAnalysis();
    } catch(err) { alert('Invalid backup file: '+err.message); }
  };
  reader.readAsText(file);
}

function escHtml(s) {
  var d = document.createElement('div');
  d.textContent = s;
  return d.innerHTML;
}

window.onload = init;
</script>
</body>
</html>"""

# Inject data
html = HTML_TEMPLATE.replace('__DATA_JSON__', DATA_JSON).replace('__INST_JSON__', INST_JSON)

out_path = r"E:\Construction APP\Marks Regester\index.html"
with open(out_path, 'w', encoding='utf-8') as f:
    f.write(html)
print(f"Created: {out_path}")
